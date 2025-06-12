from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from openpyxl.reader.excel import load_workbook
import os
import re

from diagnosis import models, personal
from .axxiliary import generate_diagnosis_advice
from .inference import analyze_image
from .batch import handle_uploaded_files, process_patient_images
from .image_processing import (
    ImagePreprocessor,
    get_processed_image_path,
    get_combined_image_path,
    combine_images,
    get_relative_path,
    parse_image_filename
)


####################
# 3. 单人诊断页面及上传处理
####################

def personal_diagnose(request):
    """
    进入单人诊断页面：
      - last_patient 显示最新添加的患者信息
      - record 显示该患者最新一条记录（按 diagnosis_time 倒序取）
      同时将 MEDIA_URL 添加到上下文中以便正确拼接图片路径。
    """
    last_patient = models.Patient.objects.last()
    record = None
    advice = None
    if last_patient:
        record = models.Record.objects.filter(patient=last_patient).order_by('-diagnosis_time').first()
        if record:
            # 如果记录为待诊断状态，则调用诊断
            if record.result == "待诊断":
                try:
                    # 采用该患者上传的原始图片进行分析
                    patient_images = models.PatientImage.objects.filter(patient=last_patient)
                    results = []
                    for patient_image in patient_images:
                        image_path = patient_image.image.path
                        analysis_result = analyze_image(image_path)
                        if isinstance(analysis_result, dict) and 'diagnosis' in analysis_result:
                            results.append(analysis_result['diagnosis'])
                    if results:
                        # 此处简单取第一个分析结果，或根据需求合并多个结果
                        record.result = results[0]
                        record.save()
                        advice_text = generate_diagnosis_advice(last_patient, record)
                        record.advice = advice_text
                        record.save()
                except Exception as e:
                    print(f"诊断过程中出错: {str(e)}")
            advice = record.advice

    context = {
        "last_patient": last_patient,
        "record": record,
        "advice": advice,
        "MEDIA_URL": settings.MEDIA_URL,
    }
    return render(request, 'individual.html', context)


def personal_info(request):
    """单人诊断上传表单处理"""
    if request.method == "GET":
        form = personal.UpModelForm()
        return render(request, "personal_info.html", {'form': form})

    form = personal.UpModelForm(data=request.POST, files=request.FILES)
    response_data = {"message": ""}
    if form.is_valid():
        try:
            instance = form.save(commit=False)
            instance.save()
            uploaded_files = request.FILES.getlist('images')
            left_image_path = None
            right_image_path = None
            doctor = request.user if request.user.is_authenticated else None
            # 保存每张上传图片到 PatientImage，并进行预处理
            for image in uploaded_files:
                patient_image = models.PatientImage.objects.create(
                    patient=instance,
                    image=image
                )
                original_image_path = patient_image.image.path
                original_filename = os.path.basename(original_image_path)
                print("上传的图片文件名:", original_filename)
                patient_id, eye_type = parse_image_filename(original_filename)
                preprocessor = ImagePreprocessor()
                if not preprocessor.validate_image(original_image_path):
                    raise ValueError("无效的图像文件")
                processed_image_path = get_processed_image_path(patient_id, eye_type, original_filename)
                preprocessor.preprocess_image(original_image_path, processed_image_path)
                if eye_type == 'left':
                    left_image_path = processed_image_path
                elif eye_type == 'right':
                    right_image_path = processed_image_path

            if left_image_path and right_image_path:
                combined_image_path = get_combined_image_path(instance.id)
                combine_images(left_image_path, right_image_path, combined_image_path)
                final_image = get_relative_path(combined_image_path)
            elif left_image_path:
                final_image = get_relative_path(left_image_path)
            elif right_image_path:
                final_image = get_relative_path(right_image_path)
            else:
                final_image = None

            record = None
            if final_image:
                # 将 Windows 反斜杠替换为正斜杠（若存在）
                final_image = final_image.replace('\\', '/')
                # 保存记录时仍将处理后图片保存到 record.processed_image字段（用于展示）
                record = models.Record.objects.create(
                    patient=instance,
                    processed_image=final_image,
                    result="待诊断",
                    advice="图像已预处理完成，等待医生诊断",
                    diagnosis_time=timezone.now(),
                    doctor=doctor
                )

                # 上传后立即进行诊断：改为使用原始上传的图片进行分析
                try:
                    patient_images = models.PatientImage.objects.filter(patient=instance)
                    results = []
                    for patient_image in patient_images:
                        image_path = patient_image.image.path
                        analysis_result = analyze_image(image_path)
                        if isinstance(analysis_result, dict) and 'diagnosis' in analysis_result:
                            results.append(analysis_result['diagnosis'])
                    if results:
                        record.result = results[0]
                        record.save()
                        advice_text = generate_diagnosis_advice(instance, record)
                        record.advice = advice_text
                        record.save()
                except Exception as e:
                    print(f"上传后诊断过程中出错: {str(e)}")

            response_data["message"] = "提交成功！患者ID：{}".format(instance.id)
            response_data["status"] = "success"
            response_data["record_id"] = record.id if record else None
        except Exception as e:
            response_data["message"] = f"保存失败：{str(e)}"
            response_data["status"] = "error"
    else:
        response_data["message"] = f"表单验证失败：{form.errors}"
        response_data["status"] = "error"
    return JsonResponse(response_data)


####################
# 4. 批量诊断页面及上传处理
####################

def batch_info(request):
    """
    批量上传表单处理：
      - 上传 Excel 文件：每行数据创建一位患者记录；
      - 上传的文件夹包含各患者的影像文件，
        通过 handle_uploaded_files 保存到 media/batch_uploads/<patient.id>/，
        再通过 process_patient_images 对影像进行处理并创建 Record 记录。
    """
    if request.method == "GET":
        template_path = settings.MEDIA_URL + 'temp/batch_template.xlsx'
        return render(request, 'batch_info.html', {'template_url': template_path})

    response_data = {"status": "success", "message": ""}
    excel_file = request.FILES.get('excel_file')
    folder_files = request.FILES.getlist('folder_upload')

    if not excel_file:
        return JsonResponse({"status": "error", "message": "请上传 Excel 文件"}, status=400)

    try:
        from django.db import transaction
        with transaction.atomic():
            # 创建批次记录
            batch = models.Batch.objects.create(
                batch_name=f"批次-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                status="处理中"
            )

            wb = load_workbook(excel_file)
            ws = wb.active
            patients_created = []

            # 读取 Excel 数据并创建患者记录
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or len(row) < 3 or not any(row[:3]):
                    continue

                name, age, gender = row[:3]
                if not name or not age or not gender:
                    continue

                patient = models.Patient.objects.create(
                    name=name,
                    age=age,
                    gender=gender
                )
                patients_created.append(patient)

            # 如果上传了文件夹，处理图像文件
            if folder_files and patients_created:
                for patient in patients_created:
                    try:
                        folder_path = handle_uploaded_files(folder_files, patient)
                        record = process_patient_images(patient, folder_path,
                                                        request.user if request.user.is_authenticated else None, batch)
                        if record:
                            record.batch = batch
                            record.save()

                        # 对创建的记录立即进行诊断：使用该患者原始图片进行分析
                        if record:
                            patient_images = models.PatientImage.objects.filter(patient=patient)
                            results = []
                            for patient_image in patient_images:
                                image_path = patient_image.image.path
                                if os.path.exists(image_path):
                                    analysis_result = analyze_image(image_path)
                                    if isinstance(analysis_result, dict) and 'diagnosis' in analysis_result:
                                        results.append(analysis_result['diagnosis'])
                            if results:
                                if len(results) > 1:
                                    disease_results = [r for r in results if r != "正常"]
                                    record.result = "、".join(disease_results) if disease_results else "正常"
                                else:
                                    record.result = results[0]
                                record.save()
                                advice_text = generate_diagnosis_advice(patient, record)
                                record.advice = advice_text
                                record.save()

                                batch.processed_patients += 1
                                batch.save()
                    except Exception as e:
                        print(f"处理患者 {patient.id} 的图像时出错: {str(e)}")
                        continue

            batch.total_patients = len(patients_created)
            batch.status = "已完成" if batch.processed_patients == batch.total_patients else "处理中"
            batch.save()

            response_data["message"] = f"成功创建 {len(patients_created)} 条患者记录"
            return JsonResponse(response_data)

    except Exception as e:
        response_data.update({
            "status": "error",
            "message": f"处理失败: {str(e)}"
        })
        return JsonResponse(response_data, status=400)


def batch_diagnose(request):
    """
    进入批量诊断页面：
      - last_patient 显示最新添加的患者信息
      - record 显示所有记录中最新的一条记录（按 diagnosis_time 倒序取）
      - batches 显示所有批次信息
    """
    last_patient = models.Patient.objects.last()
    record = models.Record.objects.order_by('-diagnosis_time').first()
    batches = models.Batch.objects.all().order_by('-upload_time')
    latest_batch = batches.first() if batches.exists() else None

    # 获取最新批次的所有记录
    batch_records = []
    if latest_batch:
        batch_records = models.Record.objects.filter(batch=latest_batch).order_by('-diagnosis_time')

        # 处理所有待诊断记录
        for batch_record in batch_records:
            if batch_record.result == "待诊断":
                try:
                    # 使用患者原始图片进行分析
                    patient_images = models.PatientImage.objects.filter(patient=batch_record.patient)
                    results = []
                    for patient_image in patient_images:
                        image_path = patient_image.image.path
                        if os.path.exists(image_path):
                            analysis_result = analyze_image(image_path)
                            if isinstance(analysis_result, dict) and 'diagnosis' in analysis_result:
                                results.append(analysis_result['diagnosis'])
                    if results:
                        if len(results) > 1:
                            disease_results = [r for r in results if r != "正常"]
                            batch_record.result = "、".join(disease_results) if disease_results else "正常"
                        else:
                            batch_record.result = results[0]
                        batch_record.save()
                        advice_text = generate_diagnosis_advice(batch_record.patient, batch_record)
                        batch_record.advice = advice_text
                        batch_record.save()

                        latest_batch.processed_patients += 1
                        latest_batch.save()

                        if latest_batch.processed_patients == latest_batch.total_patients:
                            latest_batch.status = "已完成"
                            latest_batch.save()
                except Exception as e:
                    print(f"批量诊断记录 {batch_record.id} 处理出错: {str(e)}")

    context = {
        "last_patient": last_patient,
        "record": record,
        "batches": batches,
        "latest_batch": latest_batch,
        "batch_records": batch_records,
        "MEDIA_URL": settings.MEDIA_URL,
    }
    return render(request, 'batch.html', context)


def create_diagnosis_record(patient, doctor=None, batch=None):
    """
    创建诊断记录
    :param patient: 患者对象
    :param doctor: 医生对象（可选）
    :param batch: 批次对象（可选）
    :return: 创建的记录对象
    """
    # 使用患者的第一张原始图片作为诊断图像参考（仅作展示用途）
    patient_image = models.PatientImage.objects.filter(patient=patient).first()
    processed_image = patient_image.image if patient_image else ''

    record = models.Record.objects.create(
        patient=patient,
        doctor=doctor,
        batch=batch,
        processed_image=processed_image,
        result="待诊断",
        advice="正在等待诊断结果",
        diagnosis_time=timezone.now()
    )

    return record


def gemini(request):
    pass